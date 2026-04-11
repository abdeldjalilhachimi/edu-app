"""
payroll_exporter.py

Build the output Excel workbook for payroll calculation (Tab 3).

Sheet 1 ("Titulaires"):      Confirmed employees — RETSS 9%, PARTSS 25%
Sheet 2 ("Non Titulaires"):  Non-confirmed employees — RETSS 9%, PARTSS 12.5%
Sheet 3 ("Résumé"):          Summary statistics for both categories.

All monetary values stored as numeric (float) with French format "# ##0,00".
NUMCPT stored as text (@) to preserve leading zeros.
"""

import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from modules.payroll_types import PayrollOutput, PayrollStats

# French number format: "1 234 567,89"
FRENCH_NUMBER_FORMAT = "# ##0,00"

# Styling constants (same palette as other exporters)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GRAND_TOTAL_FONT = Font(bold=True)
GRAND_TOTAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Columns in the data sheets
DATA_COLUMNS = ["NUMCPT", "NOM", "PRENOM", "BRUTSS", "RETSS", "PARTSS", "NETPAI"]
MONEY_COLUMNS = {"BRUTSS", "RETSS", "PARTSS", "NETPAI"}


def _cents_to_float(cents: int) -> float:
    """Convert integer cents back to float for Excel output."""
    return cents / 100.0


def _auto_column_widths(ws) -> None:
    """Set each column width based on header length."""
    for col_cells in ws.iter_cols(min_row=1, max_row=1):
        for cell in col_cells:
            col_letter = get_column_letter(cell.column)
            header_len = len(str(cell.value)) if cell.value else 8
            ws.column_dimensions[col_letter].width = min(header_len + 6, 50)


def _style_header_row(ws, max_col: int) -> None:
    """Apply bold white-on-blue styling to the first row."""
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _format_french_payroll(cents: int) -> str:
    """
    Format integer cents as a French display string for Streamlit metrics.

    Example: 123456789 cents → "1 234 567,89"
    """
    value = cents / 100.0
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def _build_data_sheet(ws, rows: list, sheet_title: str) -> None:
    """
    Write employee rows to a worksheet with formatting and a total row.

    Parameters
    ----------
    ws : openpyxl Worksheet
    rows : list[PayrollResult]
    sheet_title : str
        Used only for context (not written to sheet).
    """
    # Write headers
    for col_idx, header in enumerate(DATA_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    _style_header_row(ws, len(DATA_COLUMNS))

    # Write data rows with formatting applied in the same pass
    right_align = Alignment(horizontal="right")
    for row_idx, emp in enumerate(rows, start=2):
        # NUMCPT — text format
        c1 = ws.cell(row=row_idx, column=1, value=emp.numcpt_raw)
        c1.number_format = "@"
        # NOM, PRENOM — no special format
        ws.cell(row=row_idx, column=2, value=emp.nom)
        ws.cell(row=row_idx, column=3, value=emp.prenom)
        # Money columns — French number format
        for col_idx, cents in ((4, emp.brutss_cents), (5, emp.retss_cents),
                               (6, emp.partss_cents), (7, emp.netpai_cents)):
            cell = ws.cell(row=row_idx, column=col_idx, value=_cents_to_float(cents))
            cell.number_format = FRENCH_NUMBER_FORMAT
            cell.alignment = right_align

    # Total row
    if rows:
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = GRAND_TOTAL_FONT
        ws.cell(row=total_row, column=1).fill = GRAND_TOTAL_FILL

        for col_idx in range(2, 4):  # NOM, PRENOM — empty in total row
            ws.cell(row=total_row, column=col_idx).fill = GRAND_TOTAL_FILL

        totals = {
            4: sum(r.brutss_cents for r in rows),
            5: sum(r.retss_cents for r in rows),
            6: sum(r.partss_cents for r in rows),
            7: sum(r.netpai_cents for r in rows),
        }
        for col_idx, total_cents in totals.items():
            cell = ws.cell(row=total_row, column=col_idx, value=_cents_to_float(total_cents))
            cell.number_format = FRENCH_NUMBER_FORMAT
            cell.font = GRAND_TOTAL_FONT
            cell.fill = GRAND_TOTAL_FILL
            cell.alignment = right_align

    _auto_column_widths(ws)


def _build_summary_sheet(ws, stats: PayrollStats) -> None:
    """Build Sheet 3 — Summary statistics."""
    current_row = 1

    # ── Section: Counts ─────────────────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="RÉSUMÉ CALCUL PAIE")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=2)
    current_row += 1

    ws.cell(row=current_row, column=1, value="Total employés")
    ws.cell(row=current_row, column=2, value=stats.total_employees)
    current_row += 1
    ws.cell(row=current_row, column=1, value="Titulaires")
    ws.cell(row=current_row, column=2, value=stats.confirmed_count)
    current_row += 1
    ws.cell(row=current_row, column=1, value="Non Titulaires")
    ws.cell(row=current_row, column=2, value=stats.non_confirmed_count)
    current_row += 2  # blank separator

    # ── Section: Confirmed totals ───────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="TITULAIRES (PARTSS 25%)")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=2)
    current_row += 1

    for label, cents in [
        ("Total BRUTSS", stats.confirmed_brutss_cents),
        ("Total RETSS (9%)", stats.confirmed_retss_cents),
        ("Total PARTSS (25%)", stats.confirmed_partss_cents),
        ("Total NETPAI", stats.confirmed_netpai_cents),
    ]:
        ws.cell(row=current_row, column=1, value=label)
        cell = ws.cell(row=current_row, column=2, value=_cents_to_float(cents))
        cell.number_format = FRENCH_NUMBER_FORMAT
        cell.alignment = Alignment(horizontal="right")
        current_row += 1

    current_row += 1  # blank separator

    # ── Section: Non-confirmed totals ───────────────────────────────────────
    ws.cell(row=current_row, column=1, value="NON TITULAIRES (PARTSS 12,5%)")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=2)
    current_row += 1

    for label, cents in [
        ("Total BRUTSS", stats.non_confirmed_brutss_cents),
        ("Total RETSS (9%)", stats.non_confirmed_retss_cents),
        ("Total PARTSS (12,5%)", stats.non_confirmed_partss_cents),
        ("Total NETPAI", stats.non_confirmed_netpai_cents),
    ]:
        ws.cell(row=current_row, column=1, value=label)
        cell = ws.cell(row=current_row, column=2, value=_cents_to_float(cents))
        cell.number_format = FRENCH_NUMBER_FORMAT
        cell.alignment = Alignment(horizontal="right")
        current_row += 1

    current_row += 1  # blank separator

    # ── Section: Grand totals ───────────────────────────────────────────────
    ws.cell(row=current_row, column=1, value="TOTAUX GÉNÉRAUX")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=2)
    current_row += 1

    for label, cents in [
        ("Total BRUTSS", stats.grand_brutss_cents),
        ("Total RETSS", stats.grand_retss_cents),
        ("Total PARTSS", stats.grand_partss_cents),
        ("Total NETPAI", stats.grand_netpai_cents),
    ]:
        label_cell = ws.cell(row=current_row, column=1, value=label)
        label_cell.font = GRAND_TOTAL_FONT
        label_cell.fill = GRAND_TOTAL_FILL
        value_cell = ws.cell(row=current_row, column=2, value=_cents_to_float(cents))
        value_cell.number_format = FRENCH_NUMBER_FORMAT
        value_cell.font = GRAND_TOTAL_FONT
        value_cell.fill = GRAND_TOTAL_FILL
        value_cell.alignment = Alignment(horizontal="right")
        current_row += 1

    _auto_column_widths(ws)


def _build_empty_brutss_sheet(ws, empty_brutss: list) -> None:
    """
    Build a sheet listing employees whose BRUTSS was empty (set to 0).

    Parameters
    ----------
    ws : openpyxl Worksheet
    empty_brutss : list[tuple]
        [(numcpt, nom, prenom), ...]
    """
    WARNING_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    # Section header
    ws.cell(row=1, column=1, value=f"EMPLOYÉS AVEC BRUTSS VIDE ({len(empty_brutss)} trouvé(s))")
    ws.cell(row=1, column=1).font = SECTION_FONT
    ws.cell(row=1, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # Column headers
    for col_idx, header in enumerate(["NUMCPT", "NOM", "PRENOM"], start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = WARNING_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (numcpt, nom, prenom) in enumerate(empty_brutss, start=3):
        ws.cell(row=row_idx, column=1, value=numcpt)
        ws.cell(row=row_idx, column=1).number_format = "@"
        ws.cell(row=row_idx, column=2, value=nom)
        ws.cell(row=row_idx, column=3, value=prenom)

    _auto_column_widths(ws)


def create_payroll_excel(result: PayrollOutput, empty_brutss: list = None) -> bytes:
    """
    Build the 3-sheet output workbook for payroll calculation.

    Sheet 1 ("Titulaires"):      Confirmed employees with PARTSS 25%
    Sheet 2 ("Non Titulaires"):  Non-confirmed employees with PARTSS 12.5%
    Sheet 3 ("Résumé"):          Summary statistics
    Sheet 4 ("BRUTSS Vides"):    Employees with empty BRUTSS (only if any)

    Parameters
    ----------
    result : PayrollOutput
        Output from calculate_payroll().
    empty_brutss : list[tuple], optional
        [(numcpt, nom, prenom), ...] for employees with empty BRUTSS.

    Returns
    -------
    bytes
        Excel file content ready for download.
    """
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Create a dummy sheet so openpyxl has something to start with
        pd.DataFrame().to_excel(writer, sheet_name="_tmp", index=False)

        wb = writer.book

        # Sheet 1 — Titulaires
        ws1 = wb.create_sheet(title="Titulaires")
        _build_data_sheet(ws1, result.confirmed_rows, "Titulaires")

        # Sheet 2 — Non Titulaires
        ws2 = wb.create_sheet(title="Non Titulaires")
        _build_data_sheet(ws2, result.non_confirmed_rows, "Non Titulaires")

        # Sheet 3 — Résumé
        ws3 = wb.create_sheet(title="Résumé")
        _build_summary_sheet(ws3, result.stats)

        # Sheet 4 — Empty BRUTSS (only if any were found)
        if empty_brutss:
            ws4 = wb.create_sheet(title="BRUTSS Vides")
            _build_empty_brutss_sheet(ws4, empty_brutss)

        # Remove the dummy sheet
        del wb["_tmp"]

    buffer.seek(0)
    return buffer.getvalue()
