"""
checker.py

Empty-value checker for uploaded Excel files.

Scans every sheet of an uploaded .xlsx file and detects rows where one or
more of the target columns are empty (NaN, blank string, or whitespace).

Target columns (checked only if present in the sheet):
    DATNAIS, DATENT, BRUTSS, RETSS, PARTSS, NIN

Output: an Excel report with
  - a detail sheet (one row per problematic employee row, empty cells in red),
  - a summary sheet,
  - one sheet per checked column listing the records (sheet + line) where
    that specific column is empty.
"""

import io
from typing import NamedTuple

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Columns to check for empty values (only those present in each sheet are checked)
CHECK_COLUMNS = ["NIN", "DATNAIS", "DATENT", "BRUTSS", "RETSS", "PARTSS"]

# Identity columns copied into the report when present
IDENTITY_COLUMNS = ["NUMCPT", "NOM", "PRENOM"]


class CheckResult(NamedTuple):
    """Full output of run_check, consumed by the UI and the exporter."""
    report_df: pd.DataFrame   # one row per problematic source row
    per_column: dict          # {checked_col: DataFrame[FEUILLE, LIGNE, NUMCPT, NOM, PRENOM]}
    stats: dict               # summary counters for the UI


def _is_empty(series: pd.Series) -> pd.Series:
    """Boolean mask: True where the value is NaN, empty or whitespace-only."""
    as_str = series.astype(str).str.strip().str.lower()
    return series.isna() | as_str.isin(["", "nan", "none", "nat"])


def check_sheet(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """
    Check one sheet's DataFrame for empty values in the target columns.

    Returns a report DataFrame (possibly empty) with columns:
    FEUILLE, LIGNE, NUMCPT, NOM, PRENOM, <checked cols>, COLONNES_VIDES
    """
    present_checks = [c for c in CHECK_COLUMNS if c in df.columns]
    if not present_checks or df.empty:
        return pd.DataFrame()

    empty_masks = {col: _is_empty(df[col]) for col in present_checks}
    any_empty = pd.concat(empty_masks.values(), axis=1).any(axis=1)

    if not any_empty.any():
        return pd.DataFrame()

    bad = df[any_empty].copy()

    report = pd.DataFrame(index=bad.index)
    report["FEUILLE"] = sheet_name
    # +2: Excel rows are 1-based and row 1 is the header
    report["LIGNE"] = bad.index + 2

    for col in IDENTITY_COLUMNS:
        report[col] = bad[col] if col in bad.columns else ""

    for col in present_checks:
        report[col] = bad[col]

    report["COLONNES_VIDES"] = [
        ", ".join(col for col in present_checks if empty_masks[col].loc[idx])
        for idx in bad.index
    ]

    return report.reset_index(drop=True)


def run_check(file_obj, filename: str) -> CheckResult:
    """
    Scan every sheet of an Excel file for empty target values.

    Sheets without any of the target columns are skipped.

    Raises
    ------
    ValueError
        If the file cannot be read, or no sheet contains any target column.
    """
    file_obj.seek(0)
    try:
        sheets = pd.read_excel(
            file_obj, sheet_name=None, engine="openpyxl",
            dtype={"NUMCPT": str, "NIN": str, "NUMSS": str},
        )
    except Exception as e:
        raise ValueError(
            f"Cannot read file '{filename}': {e}\n"
            "Please make sure it is a valid, unencrypted .xlsx file."
        ) from e

    reports = []
    sheets_checked = 0
    sheets_skipped = 0
    rows_scanned = 0
    empty_counts = {col: 0 for col in CHECK_COLUMNS}
    columns_seen = set()  # which checked columns exist in at least one sheet

    for sheet_name, df in sheets.items():
        df.columns = df.columns.astype(str).str.strip()
        # Keep the original index so LIGNE maps back to the Excel row number
        df = df.dropna(how="all")

        # Skip grand-total rows (no NOM/PRENOM/NUMCPT) — they are not employees
        id_cols = [c for c in IDENTITY_COLUMNS if c in df.columns]
        if id_cols:
            all_id_empty = pd.concat(
                [_is_empty(df[c]) for c in id_cols], axis=1
            ).all(axis=1)
            df = df[~all_id_empty]

        present_checks = [c for c in CHECK_COLUMNS if c in df.columns]
        if not present_checks:
            sheets_skipped += 1
            continue

        columns_seen.update(present_checks)
        sheets_checked += 1
        rows_scanned += len(df)

        for col in present_checks:
            empty_counts[col] += int(_is_empty(df[col]).sum())

        report = check_sheet(df, sheet_name)
        if not report.empty:
            reports.append(report)

    if sheets_checked == 0:
        raise ValueError(
            f"File '{filename}': no sheet contains any of the checked columns "
            f"({', '.join(CHECK_COLUMNS)})."
        )

    if reports:
        report_df = pd.concat(reports, ignore_index=True)
        # Fixed column order regardless of which sheets contributed which columns
        ordered = (
            ["FEUILLE", "LIGNE"]
            + IDENTITY_COLUMNS
            + [c for c in CHECK_COLUMNS if c in report_df.columns]
            + ["COLONNES_VIDES"]
        )
        report_df = report_df.reindex(columns=ordered)
    else:
        report_df = pd.DataFrame()

    # Per-column breakdown: for each checked column, the records where it is empty
    # (FEUILLE + LIGNE + identity), so the user gets a dedicated list per column.
    per_column = _build_per_column(report_df)

    stats = {
        "sheets_checked": sheets_checked,
        "sheets_skipped": sheets_skipped,
        "rows_scanned": rows_scanned,
        "rows_with_empty": len(report_df),
        "empty_counts": {c: n for c, n in empty_counts.items() if n > 0},
        # Requested columns that exist in no sheet at all (e.g. NIN absent)
        "columns_absent": [c for c in CHECK_COLUMNS if c not in columns_seen],
    }

    return CheckResult(report_df=report_df, per_column=per_column, stats=stats)


def _build_per_column(report_df: pd.DataFrame) -> dict:
    """
    Split the flat report into one list per checked column.

    Returns {col: DataFrame[FEUILLE, LIGNE, NUMCPT, NOM, PRENOM]} containing
    only the records where that specific column is empty. Columns with no
    empties are omitted. Order follows CHECK_COLUMNS.
    """
    per_column: dict = {}
    if report_df.empty:
        return per_column

    id_present = [c for c in IDENTITY_COLUMNS if c in report_df.columns]
    base_cols = ["FEUILLE", "LIGNE"] + id_present

    vides = report_df["COLONNES_VIDES"].fillna("").astype(str)
    for col in CHECK_COLUMNS:
        # A record belongs to this column's list if `col` appears in COLONNES_VIDES
        mask = vides.str.split(", ").apply(lambda parts: col in parts)
        if mask.any():
            sub = report_df.loc[mask, base_cols].reset_index(drop=True)
            sub.index = sub.index + 1  # 1-based numbering for display
            per_column[col] = sub

    return per_column


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_EMPTY_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _write_simple_sheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to a worksheet with a styled header row."""
    columns = df.columns.tolist()
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df[col_name].head(200)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    if not df.empty:
        ws.auto_filter.ref = ws.dimensions


def create_check_excel(result: CheckResult, filename: str) -> bytes:
    """
    Build the downloadable Excel report.

    Sheet 1 ("Valeurs Vides"): one row per problematic source row,
    with the empty cells highlighted in red.
    Sheet 2 ("Résumé"):        summary statistics.
    One sheet per checked column ("Vides NIN", "Vides DATNAIS", …):
    the records (FEUILLE + LIGNE + identity) where that column is empty.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1 — detail ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Valeurs Vides"

    df = result.report_df

    if df.empty:
        ws.cell(row=1, column=1, value="Aucune valeur vide détectée ✔").font = Font(bold=True)
    else:
        columns = df.columns.tolist()
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        check_col_positions = {
            col: columns.index(col) + 1 for col in CHECK_COLUMNS if col in columns
        }

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            empty_cols = set(
                c.strip() for c in str(getattr(row, "COLONNES_VIDES", "")).split(",")
            )
            for col_idx, (col_name, value) in enumerate(zip(columns, row), start=1):
                if pd.isna(value):
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_name in check_col_positions and col_name in empty_cols:
                    cell.fill = _EMPTY_FILL

        # Column widths
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = max(
                [len(str(col_name))]
                + [len(str(v)) for v in df[col_name].head(200)]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2 — summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Résumé")
    summary_rows = [
        ("Fichier analysé", filename),
        ("Feuilles vérifiées", result.stats["sheets_checked"]),
        ("Feuilles ignorées (colonnes absentes)", result.stats["sheets_skipped"]),
        ("Lignes analysées", result.stats["rows_scanned"]),
        ("Lignes avec valeur(s) vide(s)", result.stats["rows_with_empty"]),
    ]
    for col, n in result.stats["empty_counts"].items():
        summary_rows.append((f"Cellules vides — {col}", n))

    absent = result.stats.get("columns_absent", [])
    if absent:
        summary_rows.append(("Colonnes absentes du fichier", ", ".join(absent)))

    for row_idx, (label, value) in enumerate(summary_rows, start=1):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=value)
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 30

    # ── One sheet per checked column — records where that column is empty ──
    for col, sub in result.per_column.items():
        ws_col = wb.create_sheet(title=f"Vides {col}"[:31])
        _write_simple_sheet(ws_col, sub.reset_index(drop=True))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
