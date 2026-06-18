"""
annual_exporter.py

Build the output Excel workbook for the annual declaration.

Sheet 1 ("Déclaration Annuelle"): One row per employee with quarterly BRUTSS + annual total.
Sheet 2 ("Résumé"):               Summary statistics.
Sheet 3 ("Absences"):             Employees missing from specific quarters.
Sheet 4 ("BRUTSS Vides"):         Employees with empty BRUTSS (only if any found).

Same styling palette as other exporters for consistency.
"""

import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from modules.annual_types import AnnualResult, MissingEmployee

# French number format: "1 234 567,89"
FRENCH_NUMBER_FORMAT = "# ##0,00"

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GRAND_TOTAL_FONT = Font(bold=True)
GRAND_TOTAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def _cents_to_float(cents: int) -> float:
    """Convert integer cents back to float for Excel output."""
    return cents / 100.0


def _format_french_annual(cents: int) -> str:
    """Format integer cents as a French display string for Streamlit metrics."""
    value = cents / 100.0
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def _auto_column_widths(ws) -> None:
    """Set each column width based on header length."""
    for col_cells in ws.iter_cols(min_row=1, max_row=1):
        for cell in col_cells:
            col_letter = get_column_letter(cell.column)
            header_len = len(str(cell.value)) if cell.value else 8
            ws.column_dimensions[col_letter].width = min(header_len + 6, 50)


def _style_header_row(ws, max_col: int) -> None:
    """Apply bold white-on-blue styling to the first row."""
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _build_summary_sheet(ws, stats) -> None:
    """Build Sheet 2 — Summary statistics."""
    file_labels = stats.file_labels
    quarterly_totals = stats.quarterly_totals_cents
    current_row = 1

    # Section header
    ws.cell(row=current_row, column=1, value="RÉSUMÉ DÉCLARATION ANNUELLE")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=2)
    current_row += 1

    ws.cell(row=current_row, column=1, value="Employés uniques")
    ws.cell(row=current_row, column=2, value=int(stats.unique_count))
    current_row += 2  # blank separator

    # Quarterly totals
    ws.cell(row=current_row, column=1, value="TOTAUX PAR TRIMESTRE")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=2)
    current_row += 1

    for label, total_cents in zip(file_labels, quarterly_totals):
        ws.cell(row=current_row, column=1, value=f"Total BRUTSS — {label}")
        val_cell = ws.cell(row=current_row, column=2, value=_cents_to_float(total_cents))
        val_cell.number_format = FRENCH_NUMBER_FORMAT
        val_cell.alignment = Alignment(horizontal="right")
        current_row += 1

    # Grand total
    label_cell = ws.cell(row=current_row, column=1, value="Total BRUTSS Annuel")
    label_cell.font = GRAND_TOTAL_FONT
    label_cell.fill = GRAND_TOTAL_FILL
    value_cell = ws.cell(row=current_row, column=2,
                         value=_cents_to_float(stats.grand_total_cents))
    value_cell.number_format = FRENCH_NUMBER_FORMAT
    value_cell.font = GRAND_TOTAL_FONT
    value_cell.fill = GRAND_TOTAL_FILL
    value_cell.alignment = Alignment(horizontal="right")

    _auto_column_widths(ws)


def _build_absences_sheet(ws, missing_per_file: tuple, file_labels: tuple) -> None:
    """Build the Absences sheet — employees missing from each quarter."""
    ABSENT_HEADER_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    current_row = 1

    for label, missing_list in zip(file_labels, missing_per_file):
        count = len(missing_list)
        ws.cell(row=current_row, column=1,
                value=f"ABSENTS DANS {label}  —  {count} employé(s)")
        ws.cell(row=current_row, column=1).font = SECTION_FONT
        ws.cell(row=current_row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=3)
        current_row += 1

        if not missing_list:
            ws.cell(row=current_row, column=1,
                    value="Aucun employé absent — tous présents dans ce trimestre.")
            ws.cell(row=current_row, column=1).font = Font(italic=True, color="548235")
            current_row += 2
            continue

        for col_idx, header in enumerate(["NUMCPT", "NOM", "PRENOM"], start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = ABSENT_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        for emp in missing_list:
            ws.cell(row=current_row, column=1, value=emp.numcpt_raw)
            ws.cell(row=current_row, column=1).number_format = "@"
            ws.cell(row=current_row, column=2, value=emp.nom)
            ws.cell(row=current_row, column=3, value=emp.prenom)
            current_row += 1

        current_row += 1

    _auto_column_widths(ws)


def _build_empty_brutss_sheet(ws, empty_brutss_per_file: dict) -> None:
    """Build a sheet listing employees whose BRUTSS_TOTAL was empty."""
    WARNING_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    current_row = 1

    for file_label, employees in empty_brutss_per_file.items():
        if not employees:
            continue

        ws.cell(row=current_row, column=1,
                value=f"BRUTSS VIDE DANS {file_label}  —  {len(employees)} employé(s)")
        ws.cell(row=current_row, column=1).font = SECTION_FONT
        ws.cell(row=current_row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=3)
        current_row += 1

        for col_idx, header in enumerate(["NUMCPT", "NOM", "PRENOM"], start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = WARNING_FILL
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        for numcpt, nom, prenom in employees:
            ws.cell(row=current_row, column=1, value=numcpt)
            ws.cell(row=current_row, column=1).number_format = "@"
            ws.cell(row=current_row, column=2, value=nom)
            ws.cell(row=current_row, column=3, value=prenom)
            current_row += 1

        current_row += 1

    _auto_column_widths(ws)


def create_annual_excel(
    result: AnnualResult,
    anref_year: int = 2025,
    empty_brutss_per_file: dict = None,
) -> bytes:
    """
    Assemble the output workbook for annual declaration.

    Sheet 1 ("Déclaration Annuelle"):
        NEMPLOYEUR, ANREF, N, NUMCPT, NUMSS, ADM, NOM, PRENOM,
        DATNAIS, NBRTRAV, DATENT, DATSOR,
        BRUTSS_<Q1>, BRUTSS_<Q2>, BRUTSS_<Q3>, BRUTSS_<Q4>, BRUTSS_ANNUEL,
        UNBRTRAV, OBSERV

    Sheet 2 ("Résumé"): Summary statistics.
    Sheet 3 ("Absences"): Missing employees per quarter.
    Sheet 4 ("BRUTSS Vides"): Only if any empty BRUTSS found.

    Returns
    -------
    bytes
        Excel file content ready for download.
    """
    buffer = io.BytesIO()

    file_labels = result.stats.file_labels
    brutss_col_names = [f"BRUTSS_{label}" for label in file_labels]

    # Build DataFrame
    data = []
    for row in result.rows:
        row_dict = {
            "NUMCPT": row.numcpt_raw,
            "NUMSS": row.numss,
            "ADM": row.adm,
            "NOM": row.nom,
            "PRENOM": row.prenom,
            "DATNAIS": row.datnais,
            "NBRTRAV": row.nbrtrav,
            "DATENT": row.datent,
            "DATSOR": row.datsor,
        }
        # Per-quarter days worked (needed by the official fixed-width TXT export)
        for i, label in enumerate(file_labels):
            row_dict[f"NBRTRAV_{label}"] = row.quarterly_nbrtrav[i]
        for i, col_name in enumerate(brutss_col_names):
            row_dict[col_name] = _cents_to_float(row.quarterly_brutss_cents[i])
        row_dict["BRUTSS_ANNUEL"] = _cents_to_float(row.brutss_annual)
        data.append(row_dict)

    export_df = pd.DataFrame(data)

    # Add extra columns
    export_df.insert(0, "NEMPLOYEUR", "0840198947")
    export_df.insert(1, "ANREF", int(anref_year))
    export_df.insert(2, "N", range(1, len(export_df) + 1))
    export_df["UNBRTRAV"] = "J"
    export_df["OBSERV"] = ""

    # Ensure ID columns stay as string
    if not export_df.empty:
        for col in ("NUMCPT", "NUMSS", "ADM", "NEMPLOYEUR"):
            export_df[col] = (
                export_df[col].astype(str).str.strip()
                .replace({"nan": "", "None": ""})
            )

    # All BRUTSS columns for formatting
    all_brutss_columns = set(brutss_col_names + ["BRUTSS_ANNUEL"])

    # Single-save
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Déclaration Annuelle", index=False)

        wb = writer.book
        ws1 = wb["Déclaration Annuelle"]

        _style_header_row(ws1, ws1.max_column)

        # Pre-compute column roles for single-pass formatting
        brutss_col_indices = set()
        text_col_indices = set()
        for col_idx in range(1, ws1.max_column + 1):
            header = ws1.cell(row=1, column=col_idx).value
            if header in all_brutss_columns:
                brutss_col_indices.add(col_idx)
            elif header in ("NUMCPT", "NUMSS", "NEMPLOYEUR"):
                text_col_indices.add(col_idx)

        # Single pass formatting
        right_align = Alignment(horizontal="right")
        if brutss_col_indices or text_col_indices:
            for row_idx in range(2, ws1.max_row + 1):
                for col_idx in brutss_col_indices:
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell.number_format = FRENCH_NUMBER_FORMAT
                    cell.alignment = right_align
                for col_idx in text_col_indices:
                    ws1.cell(row=row_idx, column=col_idx).number_format = "@"

        _auto_column_widths(ws1)

        # Sheet 2 — Summary
        ws2 = wb.create_sheet(title="Résumé")
        _build_summary_sheet(ws2, result.stats)

        # Sheet 3 — Absences
        ws3 = wb.create_sheet(title="Absences")
        _build_absences_sheet(ws3, result.missing_per_file, result.stats.file_labels)

        # Sheet 4 — Empty BRUTSS (only if any)
        if empty_brutss_per_file:
            total_empty = sum(len(v) for v in empty_brutss_per_file.values())
            if total_empty > 0:
                ws4 = wb.create_sheet(title="BRUTSS Vides")
                _build_empty_brutss_sheet(ws4, empty_brutss_per_file)

    buffer.seek(0)
    return buffer.getvalue()
