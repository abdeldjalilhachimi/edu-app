"""
comparator.py

Compare a secondary Excel file against an official (reliable) one, keyed on
NUMSS. The official file is treated as the source of truth.

The two files may have their header on any row (the official CNAS export keeps
a title banner above the real header), so the header row is auto-detected by
locating the row that contains the key column.

Comparison produces:
  - Records present in the official file but MISSING from the secondary file.
  - Records present in the secondary file but ABSENT from the official file
    (extra / unknown).
  - Records matched on NUMSS whose shared fields (NOM, PRENOM, DATNAIS) DIFFER
    between the two files (with date- and case-insensitive normalization to
    avoid false positives).

Output: a multi-sheet Excel report ready to download.
"""

import io
from typing import NamedTuple

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

KEY_COLUMN = "NUMSS"

# Shared fields compared when a NUMSS matches in both files
COMPARE_FIELDS = ["NOM", "PRENOM", "DATNAIS"]

# Date-like fields get parsed before comparison (mixed dd/mm/yyyy and ISO formats)
DATE_FIELDS = {"DATNAIS", "DATENT"}

# How many top rows to scan when auto-detecting the header row
HEADER_SCAN_ROWS = 30


class CompareResult(NamedTuple):
    """Full output of compare_files, consumed by the UI and the exporter."""
    missing_df: pd.DataFrame   # in official, absent from secondary
    extra_df: pd.DataFrame     # in secondary, absent from official
    diff_df: pd.DataFrame      # matched NUMSS with differing shared fields
    stats: dict


# ─────────────────────────────────────────────────────────────────────────────
# Loading / normalization
# ─────────────────────────────────────────────────────────────────────────────

def load_keyed_file(file_obj, filename: str, key: str = KEY_COLUMN) -> pd.DataFrame:
    """
    Read an Excel file whose header row contains `key`, auto-detecting that row.

    Returns a DataFrame with stripped column names and a 1-based source row
    number column (_LIGNE) pointing back to the original Excel row.

    Raises
    ------
    ValueError
        If the file can't be read or no row contains the key column.
    """
    file_obj.seek(0)
    try:
        raw = pd.read_excel(file_obj, engine="openpyxl", header=None, dtype=str)
    except Exception as e:
        raise ValueError(
            f"Cannot read file '{filename}': {e}\n"
            "Please make sure it is a valid, unencrypted .xlsx file."
        ) from e

    header_row = None
    limit = min(HEADER_SCAN_ROWS, len(raw))
    for i in range(limit):
        values = [
            str(v).strip() if v is not None else ""
            for v in raw.iloc[i].tolist()
        ]
        if key in values:
            header_row = i
            break

    if header_row is None:
        raise ValueError(
            f"File '{filename}': could not find a header row containing "
            f"the column '{key}' (scanned the first {limit} rows)."
        )

    file_obj.seek(0)
    df = pd.read_excel(file_obj, engine="openpyxl", header=header_row, dtype=str)
    df.columns = df.columns.astype(str).str.strip()

    # Record the original Excel row number before dropping blanks
    # (+2: 1-based rows, and the header occupies one row)
    df["_LIGNE"] = df.index + header_row + 2

    # Drop fully-empty rows
    data_cols = [c for c in df.columns if c != "_LIGNE"]
    df = df.dropna(how="all", subset=data_cols).reset_index(drop=True)

    if key not in df.columns:
        raise ValueError(f"File '{filename}': column '{key}' not found after parsing.")

    return df


def _normalize_key(series: pd.Series) -> pd.Series:
    """Normalize NUMSS for matching: string, trimmed, no trailing .0, no leading zeros."""
    s = series.astype(str).str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)
    s = s.str.lstrip("0")
    return s.replace("", "0")


def _norm_text(value) -> str:
    """Normalize a text field for comparison: upper, trimmed, collapsed spaces."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip().upper()
    if s.lower() in ("nan", "none", "nat"):
        return ""
    return " ".join(s.split())


def _norm_date(value) -> str:
    """Normalize a date field to YYYY-MM-DD; fall back to text if unparseable."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip()
    if s == "" or s.lower() in ("nan", "none", "nat"):
        return ""
    # ISO-like strings (YYYY-MM-DD…) are month-first; everything else is day-first
    iso_like = len(s) >= 4 and s[:4].isdigit() and ("-" in s or ":" in s)
    parsed = pd.to_datetime(s, dayfirst=not iso_like, errors="coerce")
    if pd.isna(parsed):
        return _norm_text(value)
    return parsed.strftime("%Y-%m-%d")


def _norm_field(column: str, value) -> str:
    """Dispatch to the date or text normalizer based on the column name."""
    if column in DATE_FIELDS:
        return _norm_date(value)
    return _norm_text(value)


# ─────────────────────────────────────────────────────────────────────────────
# Comparison
# ─────────────────────────────────────────────────────────────────────────────

def compare_files(
    official_obj, official_name: str,
    secondary_obj, secondary_name: str,
    key: str = KEY_COLUMN,
) -> CompareResult:
    """
    Compare the secondary file against the official (reliable) one by NUMSS.

    Returns a CompareResult with the missing / extra / differing record sets.

    Raises
    ------
    ValueError
        On read errors or if either file lacks the key column.
    """
    off = load_keyed_file(official_obj, official_name, key)
    sec = load_keyed_file(secondary_obj, secondary_name, key)

    off["_KEY"] = _normalize_key(off[key])
    sec["_KEY"] = _normalize_key(sec[key])

    # Ignore blank keys for set membership (can't match on an empty NUMSS)
    off_valid = off[off["_KEY"].ne("0") & off["_KEY"].ne("")]
    sec_valid = sec[sec["_KEY"].ne("0") & sec["_KEY"].ne("")]

    off_keys = set(off_valid["_KEY"])
    sec_keys = set(sec_valid["_KEY"])

    missing_keys = off_keys - sec_keys   # official-only
    extra_keys = sec_keys - off_keys     # secondary-only
    common_keys = off_keys & sec_keys

    # ── Missing: official records absent from secondary ──────────────────
    missing_df = (
        off_valid[off_valid["_KEY"].isin(missing_keys)]
        .drop(columns=["_KEY"])
        .reset_index(drop=True)
    )
    missing_df = _front_columns(missing_df, key)

    # ── Extra: secondary records absent from official ────────────────────
    extra_df = (
        sec_valid[sec_valid["_KEY"].isin(extra_keys)]
        .drop(columns=["_KEY"])
        .reset_index(drop=True)
    )
    extra_df = _front_columns(extra_df, key)

    # ── Differences on shared fields for matched NUMSS ───────────────────
    diff_df = _build_diff(off_valid, sec_valid, common_keys, key)

    stats = {
        "official_name": official_name,
        "secondary_name": secondary_name,
        "official_rows": len(off),
        "secondary_rows": len(sec),
        "official_keys": len(off_keys),
        "secondary_keys": len(sec_keys),
        "matched": len(common_keys),
        "missing_count": len(missing_keys),
        "extra_count": len(extra_keys),
        "diff_count": len(diff_df),
        "official_blank_keys": int((~off.index.isin(off_valid.index)).sum()),
        "secondary_blank_keys": int((~sec.index.isin(sec_valid.index)).sum()),
    }

    return CompareResult(
        missing_df=missing_df,
        extra_df=extra_df,
        diff_df=diff_df,
        stats=stats,
    )


def _front_columns(df: pd.DataFrame, key: str) -> pd.DataFrame:
    """Reorder so _LIGNE and the key column come first; keep the rest in order."""
    if df.empty:
        return df
    front = [c for c in ["_LIGNE", key] if c in df.columns]
    rest = [c for c in df.columns if c not in front]
    df = df[front + rest].rename(columns={"_LIGNE": "LIGNE"})
    return df


def _build_diff(off_valid, sec_valid, common_keys, key) -> pd.DataFrame:
    """
    For NUMSS present in both files, report rows where a shared field differs.

    Uses the first occurrence per key in each file. Comparison is normalized
    (dates parsed, text upper/trimmed) so only meaningful differences show up.
    """
    if not common_keys:
        return pd.DataFrame()

    fields = [f for f in COMPARE_FIELDS if f in off_valid.columns and f in sec_valid.columns]
    if not fields:
        return pd.DataFrame()

    off_first = off_valid.drop_duplicates("_KEY", keep="first").set_index("_KEY")
    sec_first = sec_valid.drop_duplicates("_KEY", keep="first").set_index("_KEY")

    rows = []
    for k in sorted(common_keys):
        o = off_first.loc[k]
        s = sec_first.loc[k]
        diffs = {}
        for f in fields:
            ov = _norm_field(f, o.get(f))
            sv = _norm_field(f, s.get(f))
            if ov != sv:
                diffs[f] = (o.get(f), s.get(f))
        if diffs:
            record = {
                key: o.get(key),
                "LIGNE_OFFICIEL": o.get("_LIGNE"),
                "LIGNE_SECONDAIRE": s.get("_LIGNE"),
            }
            for f in fields:
                if f in diffs:
                    record[f"{f}_OFFICIEL"] = diffs[f][0]
                    record[f"{f}_SECONDAIRE"] = diffs[f][1]
            record["CHAMPS_DIFFERENTS"] = ", ".join(diffs.keys())
            rows.append(record)

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, df: pd.DataFrame, empty_msg: str) -> None:
    """Write a DataFrame with a styled header, or a placeholder if empty."""
    if df.empty:
        ws.cell(row=1, column=1, value=empty_msg).font = Font(bold=True)
        return

    columns = df.columns.tolist()
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df[col_name].head(200)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def create_comparison_excel(result: CompareResult) -> bytes:
    """
    Build the downloadable comparison report.

    Sheets:
      Résumé                         — summary counters
      Manquants (officiel)           — in official, missing from secondary
      En trop (secondaire)           — in secondary, absent from official
      Différences                    — matched NUMSS with differing fields
    """
    wb = openpyxl.Workbook()
    stats = result.stats

    # ── Résumé ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Résumé"
    summary_rows = [
        ("Fichier officiel (fiable)", stats["official_name"]),
        ("Fichier secondaire (comparé)", stats["secondary_name"]),
        ("", ""),
        ("Lignes — officiel", stats["official_rows"]),
        ("Lignes — secondaire", stats["secondary_rows"]),
        ("NUMSS uniques — officiel", stats["official_keys"]),
        ("NUMSS uniques — secondaire", stats["secondary_keys"]),
        ("", ""),
        ("NUMSS correspondants", stats["matched"]),
        ("Manquants dans le secondaire", stats["missing_count"]),
        ("En trop dans le secondaire", stats["extra_count"]),
        ("Correspondances avec différences", stats["diff_count"]),
        ("", ""),
        ("NUMSS vides — officiel", stats["official_blank_keys"]),
        ("NUMSS vides — secondaire", stats["secondary_blank_keys"]),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=1):
        if label:
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 40

    # ── Detail sheets ────────────────────────────────────────────────────
    _write_sheet(
        wb.create_sheet("Manquants (officiel)"),
        result.missing_df,
        "Aucun manquant — tous les NUMSS officiels sont présents dans le secondaire ✔",
    )
    _write_sheet(
        wb.create_sheet("En trop (secondaire)"),
        result.extra_df,
        "Aucun NUMSS en trop — le secondaire ne contient que des NUMSS officiels ✔",
    )
    _write_sheet(
        wb.create_sheet("Différences"),
        result.diff_df,
        "Aucune différence de champ sur les NUMSS correspondants ✔",
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
