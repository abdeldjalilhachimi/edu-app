"""
exporter.py

Build the output Excel workbook with two sheets.

Sheet 1 ("Résultat"): Main file data with updated BRUTSS values.
Sheet 2 ("Résumé"):   Matches table + stats summary.

Strategy: use pandas ExcelWriter to write data (handles all type conversions
safely), then open with openpyxl to apply French number formatting and styling.
This avoids the "We found a problem with some content" Excel repair warning.
"""

import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# French number format: "1 234 567,89"
FRENCH_NUMBER_FORMAT = "# ##0,00"

# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GRAND_TOTAL_FONT = Font(bold=True)
GRAND_TOTAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Internal key column — excluded from all output
KEY_COLUMN = "_KEY"
BRUTSS_COLUMN = "BRUTSS"


def _auto_column_widths(ws) -> None:
    """Set each column width based on header length (fast, avoids scanning all rows)."""
    for col_cells in ws.iter_cols(min_row=1, max_row=1):
        for cell in col_cells:
            col_letter = get_column_letter(cell.column)
            header_len = len(str(cell.value)) if cell.value else 8
            ws.column_dimensions[col_letter].width = min(header_len + 6, 50)


def _style_header_row(ws, max_col: int) -> None:
    """Apply bold white-on-blue styling to the first row."""
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _apply_brutss_format(ws, brutss_col_idx: int, max_row: int) -> None:
    """Apply French number format to every BRUTSS data cell."""
    for row_idx in range(2, max_row + 1):
        cell = ws.cell(row=row_idx, column=brutss_col_idx)
        cell.number_format = FRENCH_NUMBER_FORMAT
        cell.alignment = Alignment(horizontal="right")


def build_summary_sheet(
    ws,
    duplicates: list,
    stats: dict,
) -> None:
    """
    Build Sheet 2 — Summary Report.

    Section A: Matched/added rows table
    Section B: Stats summary
    """
    current_row = 1

    # ── Section A: Matches ───────────────────────────────────────────────────

    ws.cell(row=current_row, column=1, value="CORRESPONDANCES TROUVÉES")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row, end_column=6
    )
    current_row += 1

    if not duplicates:
        ws.cell(
            row=current_row, column=1,
            value="Aucune correspondance trouvée entre le fichier principal et les fichiers additionnels."
        )
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="548235")
        current_row += 2
    else:
        ws.cell(
            row=current_row, column=1,
            value=f"{len(duplicates)} correspondance(s) trouvée(s)"
        )
        ws.cell(row=current_row, column=1).font = Font(bold=True, color="2F5496")
        current_row += 1

        # Column headers
        headers = [
            "NUMCPT", "NOM", "PRENOM",
            "BRUTSS Principal", "BRUTSS Additionnels", "BRUTSS Somme"
        ]
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Data rows
        for match in duplicates:
            ws.cell(row=current_row, column=1, value=match.numcpt)
            ws.cell(row=current_row, column=2, value=match.nom)
            ws.cell(row=current_row, column=3, value=match.prenom)

            for col_idx, val in [(4, match.brutss_main), (5, match.brutss_additional), (6, match.brutss_sum)]:
                cell = ws.cell(row=current_row, column=col_idx, value=float(val))
                cell.number_format = FRENCH_NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")

            current_row += 1

        current_row += 1  # blank separator

    # ── Section B: Stats summary ─────────────────────────────────────────────

    ws.cell(row=current_row, column=1, value="RÉSUMÉ")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row, end_column=2
    )
    current_row += 1

    stat_rows = [
        ("Lignes dans le résultat final", stats["total"]),
        ("Correspondances trouvées (dans les 2)", stats["duplicate_count"]),
        ("Nouvelles lignes ajoutées (additionnels uniquement)", stats.get("added_count", 0)),
    ]
    for label, value in stat_rows:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=int(value))
        current_row += 1

    # Grand total
    label_cell = ws.cell(row=current_row, column=1, value="Total BRUTSS final")
    label_cell.font = GRAND_TOTAL_FONT
    label_cell.fill = GRAND_TOTAL_FILL
    value_cell = ws.cell(row=current_row, column=2, value=float(stats["brutss_total"]))
    value_cell.number_format = FRENCH_NUMBER_FORMAT
    value_cell.font = GRAND_TOTAL_FONT
    value_cell.fill = GRAND_TOTAL_FILL
    value_cell.alignment = Alignment(horizontal="right")

    _auto_column_widths(ws)


def create_output_excel(
    updated_main_df: pd.DataFrame,
    duplicates: list,
    stats: dict,
    anref_year: int = 2025,
) -> bytes:
    """
    Assemble the two-sheet output workbook and return as bytes.

    Uses pandas ExcelWriter to safely write the DataFrame (handles NaN,
    numpy types, mixed dtypes), then patches formatting with openpyxl.
    """
    buffer = io.BytesIO()

    # Prepare export DataFrame: drop _KEY, keep all other columns
    export_df = updated_main_df.drop(columns=[KEY_COLUMN], errors="ignore").copy()

    # Add 5 new columns: NEMPLOYEUR, ANREF, N, UNBRTRAV, OBSERV
    export_df.insert(0, "NEMPLOYEUR", "0840198947")
    export_df.insert(1, "ANREF", int(anref_year))
    export_df.insert(2, "N", range(1, len(export_df) + 1))
    export_df["UNBRTRAV"] = "J"
    export_df["OBSERV"] = ""

    # Force large-number ID columns to string (vectorized chain)
    TEXT_COLUMNS = {"NUMCPT", "NUMSS", "NUMMUT", "MATRI", "NEMPLOYEUR"}
    for col in TEXT_COLUMNS:
        if col in export_df.columns:
            export_df[col] = (
                export_df[col].astype(str).str.strip().replace("nan", "")
            )

    # ── Single-save: write data + format + Sheet 2 in one pass ──────────────

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Résultat", index=False)

        wb = writer.book
        ws1 = wb["Résultat"]

        # Style header row
        _style_header_row(ws1, ws1.max_column)

        # Pre-compute column roles for single-pass formatting
        brutss_cols = set()
        text_cols = set()
        for col_idx in range(1, ws1.max_column + 1):
            header = ws1.cell(row=1, column=col_idx).value
            if header == BRUTSS_COLUMN:
                brutss_cols.add(col_idx)
            elif header in ("NUMCPT", "NUMSS", "NUMMUT", "MATRI"):
                text_cols.add(col_idx)

        # Single pass over all data cells
        right_align = Alignment(horizontal="right")
        if brutss_cols or text_cols:
            for row_idx in range(2, ws1.max_row + 1):
                for col_idx in brutss_cols:
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell.number_format = FRENCH_NUMBER_FORMAT
                    cell.alignment = right_align
                for col_idx in text_cols:
                    ws1.cell(row=row_idx, column=col_idx).number_format = "@"

        _auto_column_widths(ws1)

        # Build Sheet 2 — Summary
        ws2 = wb.create_sheet(title="Résumé")
        build_summary_sheet(ws2, duplicates, stats)

    buffer.seek(0)
    return buffer.getvalue()
