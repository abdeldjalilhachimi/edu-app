"""
trimestrial_exporter.py

Build the output Excel workbook for trimestrial BRUTSS consolidation.

Sheet 1 ("Consolidation"): One row per unique employee with monthly + total BRUTSS.
Sheet 2 ("Résumé"):        Summary statistics.

Column names are derived from uploaded filenames:
    NUMCPT, NOM, PRENOM, BRUTSS_<file1>, BRUTSS_<file2>, BRUTSS_<file3>, BRUTSS_TOTAL

All BRUTSS values are stored as numeric (float) in Excel cells so they can be
used in formulas. French number format "# ##0,00" is applied for display.

NUMCPT is stored as text (@) to preserve leading zeros.

Strategy: single-save via pandas ExcelWriter context (same pattern as exporter.py).
"""

import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from modules.trimestrial_types import TrimestrialResult, MissingEmployee

# French number format: "1 234 567,89"
FRENCH_NUMBER_FORMAT = "# ##0,00"

# Styling constants (same palette as main exporter for consistency)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GRAND_TOTAL_FONT = Font(bold=True)
GRAND_TOTAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def _cents_to_float(cents: int) -> float:
    """Convert integer cents back to float for Excel output."""
    return cents / 100.0


def _auto_column_widths(ws) -> None:
    """Set each column width based on header length (fast, avoids scanning all rows)."""
    for col_cells in ws.iter_cols(min_row=1, max_row=1):
        for cell in col_cells:
            col_letter = get_column_letter(cell.column)
            header_len = len(str(cell.value)) if cell.value else 8
            ws.column_dimensions[col_letter].width = min(header_len + 6, 50)


def _style_header_row(ws, max_col: int) -> None:
    """Apply bold white-on-blue styling to the first row."""
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _format_french(cents: int) -> str:
    """
    Format integer cents as a French display string for Streamlit metrics.

    Example: 123456789 cents → "1 234 567,89"
    """
    value = cents / 100.0
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def _build_brutss_column_names(file_labels: tuple) -> list:
    """
    Build the dynamic BRUTSS column names from file labels.

    Example: ("paie_jan", "paie_fev", "paie_mar")
             → ["BRUTSS_paie_jan", "BRUTSS_paie_fev", "BRUTSS_paie_mar"]

    Parameters
    ----------
    file_labels : tuple of str
        Clean labels derived from uploaded filenames (without extension).

    Returns
    -------
    list of str
        Column names for each monthly BRUTSS.
    """
    return [f"BRUTSS_{label}" for label in file_labels]


def _build_summary_sheet(ws, stats) -> None:
    """
    Build Sheet 2 — Summary statistics for trimestrial consolidation.

    Uses actual file labels for per-month total rows.
    """
    file_labels = stats.file_labels
    monthly_totals = stats.monthly_totals_cents
    current_row = 1

    # Section header
    ws.cell(row=current_row, column=1, value="RÉSUMÉ TRIMESTRIEL")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row, end_column=2,
    )
    current_row += 1

    # Unique count
    ws.cell(row=current_row, column=1, value="Employés uniques")
    ws.cell(row=current_row, column=2, value=int(stats.unique_count))
    current_row += 1

    current_row += 1  # blank separator

    # Monthly totals section
    ws.cell(row=current_row, column=1, value="TOTAUX PAR FICHIER")
    ws.cell(row=current_row, column=1).font = SECTION_FONT
    ws.cell(row=current_row, column=1).fill = SECTION_FILL
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row, end_column=2,
    )
    current_row += 1

    # One row per file, using actual filename as label
    for label, total_cents in zip(file_labels, monthly_totals):
        ws.cell(row=current_row, column=1, value=f"Total BRUTSS — {label}")
        val_cell = ws.cell(row=current_row, column=2, value=_cents_to_float(total_cents))
        val_cell.number_format = FRENCH_NUMBER_FORMAT
        val_cell.alignment = Alignment(horizontal="right")
        current_row += 1

    # Grand total with highlight
    label_cell = ws.cell(row=current_row, column=1, value="Total BRUTSS Trimestre")
    label_cell.font = GRAND_TOTAL_FONT
    label_cell.fill = GRAND_TOTAL_FILL
    value_cell = ws.cell(
        row=current_row, column=2,
        value=_cents_to_float(stats.grand_total_cents),
    )
    value_cell.number_format = FRENCH_NUMBER_FORMAT
    value_cell.font = GRAND_TOTAL_FONT
    value_cell.fill = GRAND_TOTAL_FILL
    value_cell.alignment = Alignment(horizontal="right")

    _auto_column_widths(ws)


def _build_absences_sheet(ws, missing_per_file: tuple, file_labels: tuple) -> None:
    """
    Build the "Absences" sheet — lists employees missing from each file.

    For each file, a section shows:
    - Section header with file label and count
    - Table of missing employees (NUMCPT, NOM, PRENOM)
    - Or a green message if no one is missing

    Parameters
    ----------
    ws : openpyxl Worksheet
    missing_per_file : tuple of list[MissingEmployee]
        One list per file, in the same order as file_labels.
    file_labels : tuple of str
        Clean labels derived from filenames.
    """
    # Light red fill for absent-employee rows header
    ABSENT_HEADER_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    current_row = 1

    for file_idx, (label, missing_list) in enumerate(zip(file_labels, missing_per_file)):
        # Section header: "ABSENTS DANS <filename> (N employé(s))"
        count = len(missing_list)
        section_title = f"ABSENTS DANS {label}  —  {count} employé(s)"
        ws.cell(row=current_row, column=1, value=section_title)
        ws.cell(row=current_row, column=1).font = SECTION_FONT
        ws.cell(row=current_row, column=1).fill = SECTION_FILL
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=3,
        )
        current_row += 1

        if not missing_list:
            ws.cell(
                row=current_row, column=1,
                value="Aucun employé absent — tous les employés sont présents dans ce fichier."
            )
            ws.cell(row=current_row, column=1).font = Font(italic=True, color="548235")
            current_row += 2
            continue

        # Table headers
        for col_idx, header in enumerate(["NUMCPT", "NOM", "PRENOM"], start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = ABSENT_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Data rows
        for emp in missing_list:
            ws.cell(row=current_row, column=1, value=emp.numcpt_raw)
            ws.cell(row=current_row, column=1).number_format = "@"  # text format
            ws.cell(row=current_row, column=2, value=emp.nom)
            ws.cell(row=current_row, column=3, value=emp.prenom)
            current_row += 1

        current_row += 1  # blank separator between sections

    _auto_column_widths(ws)


def create_trimestrial_excel(result: TrimestrialResult, anref_year: int = 2025) -> bytes:
    """
    Assemble the two-sheet output workbook for trimestrial consolidation.

    Sheet 1 ("Consolidation"):
        Columns: NUMCPT, NOM, PRENOM, BRUTSS_<file1>, BRUTSS_<file2>, BRUTSS_<file3>, BRUTSS_TOTAL
        - NUMCPT stored as text (@ format) to preserve leading zeros
        - BRUTSS columns stored as numeric floats with French format
        - Column names derived from uploaded filenames

    Sheet 2 ("Résumé"):
        Summary statistics (unique count, per-file totals, grand total).

    Parameters
    ----------
    result : TrimestrialResult
        Output from merge_trimestrial().

    Returns
    -------
    bytes
        Excel file content ready for download.
    """
    buffer = io.BytesIO()

    file_labels = result.stats.file_labels
    brutss_col_names = _build_brutss_column_names(file_labels)

    # ── Build DataFrame from consolidated rows ───────────────────────────────

    data = []
    for row in result.rows:
        row_dict = {
            "NUMCPT": row.numcpt_raw,
            "NUMSS": row.numss,
            "ADM": row.adm,
            "NOM": row.nom,
            "PRENOM": row.prenom,
            "DATNAIS": row.datnais,
            "NBRTRAV": row.nbrtrav,
            "DATENT": row.datent,
            "DATSOR": row.datsor,
        }
        # Add one BRUTSS column per file, named after the filename
        for i, col_name in enumerate(brutss_col_names):
            row_dict[col_name] = _cents_to_float(row.monthly_brutss_cents[i])
        row_dict["BRUTSS_TOTAL"] = _cents_to_float(row.brutss_total)
        data.append(row_dict)

    export_df = pd.DataFrame(data)

    # Add 5 extra columns: NEMPLOYEUR, ANREF, N, UNBRTRAV, OBSERV
    export_df.insert(0, "NEMPLOYEUR", "0840198947")
    export_df.insert(1, "ANREF", int(anref_year))
    export_df.insert(2, "N", range(1, len(export_df) + 1))
    export_df["UNBRTRAV"] = "J"
    export_df["OBSERV"] = ""

    # Ensure ID columns stay as string (preserve leading zeros)
    if not export_df.empty:
        for col in ("NUMCPT", "NUMSS", "ADM", "NEMPLOYEUR"):
            export_df[col] = export_df[col].astype(str).str.strip()
            export_df[col] = export_df[col].replace({"nan": "", "None": ""})

    # All BRUTSS columns that need French number formatting
    all_brutss_columns = set(brutss_col_names + ["BRUTSS_TOTAL"])

    # ── Single-save: write data + format + Sheet 2 in one pass ───────────────

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Consolidation", index=False)

        wb = writer.book
        ws1 = wb["Consolidation"]

        # Style header row
        _style_header_row(ws1, ws1.max_column)

        # Apply French number format to all BRUTSS columns (dynamic names)
        for col_idx in range(1, ws1.max_column + 1):
            header = ws1.cell(row=1, column=col_idx).value
            if header in all_brutss_columns:
                for row_idx in range(2, ws1.max_row + 1):
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell.number_format = FRENCH_NUMBER_FORMAT
                    cell.alignment = Alignment(horizontal="right")

        # Force ID columns to text format (@) to preserve leading zeros
        for col_idx in range(1, ws1.max_column + 1):
            header = ws1.cell(row=1, column=col_idx).value
            if header in ("NUMCPT", "NUMSS", "NEMPLOYEUR"):
                for row_idx in range(2, ws1.max_row + 1):
                    ws1.cell(row=row_idx, column=col_idx).number_format = "@"

        _auto_column_widths(ws1)

        # Build Sheet 2 — Summary
        ws2 = wb.create_sheet(title="Résumé")
        _build_summary_sheet(ws2, result.stats)

        # Build Sheet 3 — Absences (employees missing per file)
        ws3 = wb.create_sheet(title="Absences")
        _build_absences_sheet(ws3, result.missing_per_file, result.stats.file_labels)

    buffer.seek(0)
    return buffer.getvalue()
